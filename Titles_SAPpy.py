import pandas as pd
import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === Configuración ===
input_file = r"C:\Users\I757755\OneDrive - SAP SE\Titles\Titles Automatizados.xlsx"
output_file = r"C:\Users\I757755\OneDrive - SAP SE\Titles\Titles Automatizados_Resultado.xlsx"

# === Reemplazos de codificación comunes ===
reemplazos = {
    "Ã¡": "á", "Ã©": "é", "Ã­": "í", "Ã³": "ó", "Ãº": "ú",
    "Ã": "Á", "Ã‰": "É", "Ã": "Í", "Ã“": "Ó", "Ãš": "Ú",
    "Ã±": "ñ", "Ã": "Ñ", "Â¿": "¿", "Â¡": "¡", "Âº": "º", "Âª": "ª",
    "â€“": "–", "â€”": "—", "â€œ": "“", "â€": "”", "â€˜": "‘", "â€™": "’",
    "â€¦": "…", "Â·": "·", "Â": "", "Ãƒ³": "ó", "Ãƒ":"í"
}

# === Leer el archivo ===
df = pd.read_excel(input_file, usecols=[0, 1], names=['Key', 'Titles'], header=0)
df = df[df['Titles'].notna()].copy()

# --- Limpieza previa ---
def limpiar_texto(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto)
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    return texto


df['Titles'] = df['Titles'].apply(limpiar_texto)

# ============================================================
# 🧩 PERSONA NAME BASES
# ============================================================

persona_bases = [
    {
        "base_name": "Digital Marketing",
        "keywords": [
            "marketing digital", "seo", "sem", "ppc", "ads", "google ads", "meta ads",
            "performance marketing", "growth", "growth marketing", "content marketing",
            "email marketing", "crm marketing", "marketing automation", "hubspot",
            "marketo", "pardot", "ga4", "google analytics", "tag manager",
            "programmatic", "rtb", "social media", "community manager",
            "conversion rate", "a/b testing", "optimizacion de conversiones",
            "landing pages", "funnels", "retargeting", "remarketing", "media", "digital marketing",
            "mkt digital", "digital mkt"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "E-Commerce",
        "keywords": [
            "e-commerce", "ecommerce", "comercio electronico", "tienda online",
            "marketplace", "shopify", "magento", "woocommerce", "vtex", "prestashop",
            "salesforce commerce", "pdp", "checkout", "carrito", "merchandising",
            "catalogo", "pim", "oms", "order management", "fulfillment", "conversion",
            "cro", "a/b test", "marketplaces manager", "pricing", "promociones", "negocios",
            "digital commerce", "e commerce", "E-Billing"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | HRIT",
        "keywords": [
            "hrit", "hris", "sistemas de rrhh", "workday", "successfactors", "sap hcm",
            "sap sf", "oracle hcm", "peoplesoft", "bamboohr", "adp", "cornerstone",
            "dayforce", "kronos", "ukg", "time & attendance", "hr it", "h.r i.t", "h.r it", "hr i.t"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Compensation and Benefits",
        "keywords": [
            "compensacion", "beneficios", "comp & ben", "c&b", "rewars",
            "equidad salarial", "estructura salarial", "bandas salariales",
            "incentivos", "bonos", "remuneraciones", "benefits", "benefit", "comp",
            "esquema de comisiones", "compensaciones", "compensation", "rewards"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Payroll",
        "keywords": [
            "nomina", "payroll", "liquidacion", "remuneraciones", "planilla", "nonima", "nonimas",
            "sueldos", "salarios", "haberes", "tss", "adp", "nominas", "n?minas"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | C-Level",
        "keywords": [
            "rrhh", "rh", "hr", "humanos", "humano", "huamanos", "cultural",
            "people", "humana", "cultura", "chro", "human", "cap hum",
            "chief human resources officer", "RR. HH.", "rr hh", "rr.hh",
            "R.R.H.H.", "hhrr", "hrbp", "hcm", "h.r.", "h.r", "humans", "r.h", "r.h."
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Data Protection, Privacy & Information Security",
        "keywords": [
            "seguridad informatica", "privacidad", "ciso", "dipo", "ciberseguridad",
            "seguridad de la informacion", "data security", "cybersecurity",
            "proteccion de datos", "data protection", "privacy officer",
            "security officer", "access management", "cloud security", "cyber",
            "nube", "vulnerabilidades", "vulnerability", "data loss"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Analytics and Data Warehousing",
        "keywords": [
            "business intelligence", "inteligencia de negocios", "inteligencia de negocio",
            "inteligencia negocio", "inteligencia negocios", "bi", "data science",
            "data y analitica", "business analysis", "analiticos", "analitica",
            "analitica de informacion", "analitica de datos", "ciencia de datos",
            "cientifico de datos", "machine learning", "data analyst", "bussiness intelligence",
            "visualizacion de datos", "dashboard", "data warehouse", "analytics"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Inform. Technology",
        "keywords": [
            "it", "ti", "sistemas", "infraestructura", "tecnologia", "informatico",
            "devops", "cyberseguridad", "cloud", "digitalizacion", "software",
            "soporte tecnico", "helpdesk", "sysadmin", "system administrator",
            "mesa de ayuda", "helpdesk", "help desk", "service desk", "itsm",
            "itil", "active directory", "windows server", "linux server", "vmware",
            "esxi", "vcenter", "hyper-v", "virtualizacion", "citrix", "redes", "red",
            "networking", "switching", "routing", "dns", "dhcp", "vpn", "sd-wan",
            "wifi", "wireless", "tecnologias", "informaticos", "informatica", "chatbot",
            "cto", "cio", "T.I.", "I.T.", "info", "systems", "technology", "tech", "T. I.",
            "informacion", "information", "T.I", "programming", "system manager", "application",
            "programacion", "Applications", "computer", "tencologia", "technological", "i.t", "tecnoloogicos",
            "tecnolog?a", "programador", "tenologia", "BTP", "tecnologicas", "tecnologica"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Database, Data Management & Information",
        "keywords": [
            "datos", "data", "metadatos", "metadata", "sql", "nosql",
            "mysql", "postgresql", "postgres", "database", "architect", "internet",
            "Infrastructure", "infraestructura", "cloud", "backend", "internet", "site",
            "arquitect", "arquitecto", "informatique"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Diversity and Inclusions",
        "keywords": [
            "diversidad", "inclusion", "equidad", "dei", "de&i", "diversity",
            "belonging", "igualdad de oportunidades", "genero", "lgbt", "accesibilidad"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Employee Experience",
        "keywords": [
            "employee", "empleado", "engagement", "relations", "personas",
            "clima laboral", "bienestar", "wellbeing", "cultura", "encuestas",
            "onboarding", "offboarding", "comunicacion interna", "employer branding",
            "personal", "relaciones", "staff", "laboral", "chief experience officer"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Generalist",
        "keywords": ["FALTA COMPLETAR"],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Learning & Development",
        "keywords": [
            "l&d", "aprendizaje", "formacion", "capacitacion", "desarrollo",
            "training", "upskilling", "reskilling", "lms", "coaching", "mentoring",
            "universidad corporativa", "development"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Human Resources | Talent Acquisition",
        "keywords": [
            "talent", "reclutamiento", "seleccion", "recruiting",
            "headhunter", "sourcing", "sourcer", "bolsa de trabajo",
            "entrevista", "ta", "onboarding", "employer branding",
            "acquisition", "personas", "recruitment", "talento"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Supply Chain Planning",
        "keywords": [
            "planificacion de la cadena", "supply", "scp",
            "planificacion de demanda", "demand planning", "s&op", "snop", "ibp",
            "pronostico", "forecasting", "mrp", "mps", "drp", "aps", "replenishment",
            "inventory", "production", "produccion", "sap apo", "sap ibp", "kinaxis",
            "o9", "blue yonder", "jda", "anaplan", "e2open", "cpfr", "demand planner",
            "supply planner", "suministro", "sumnistro", "producci?n", "produccin", "automotriz"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Operations",
        "keywords": [
            "operaciones", "operational excellence", "operativa", "opex",
            "coo", "procesos", "bpm", "lean", "six sigma", "kaizen", "kanban",
            "just in time", "jit", "oee", "mejora continua", "operating","operativo",
            "optimizaciones", "optimizacion", "tecnica", "operations", "operation",
            "productivity", "automation", "operacional"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Plant Maintenance",
        "keywords": [
            "planta", "mantenimiento", "shutdown", "cmms", "sap pm", "maximo",
            "infor eam", "emaint", "plant", "confiabilidad", "reliability", "rcm",
            "tpm", "lubricacion", "instrumentacion", "reliability engineer", "line manager",
            "jefe de linea", "factory", "almacen", "almacenes", "stock"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Procurement/Purchasing",
        "keywords": [
            "compras", "procurement", "sourcing", "abastecimiento", "proveedores", "abastecimimientos",
            "negociacion", "category manager", "buyer", "comprador", "licitacion", "purchases",
            "rfq", "rfi", "rfx", "contrataciones", "purchase", "purchasing", "srm",
            "vendor management", "supplier", "ahorros", "spend", "ariba", "coupa",
            "jaggaer", "sap mm", "compra", "suministros", "cpo", "adquisiciones", "acquisitions"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Product Development / R&D / Engineering",
        "keywords": [
            "i+d", "r&d", "investigacion y desarrollo", "desarrollo de producto",
            "npi", "npd", "introduccion de producto", "ingenieria", "engineering",
            "ingeniero", "plm", "pdm", "gestion de requisitos", "prototipos",
            "laboratorio", "ensayos", "materiales", "mecatronica", "electronica",
            "mecanica", "embedded", "firmware", "hardware", "diseno de experimentos",
            "doe", "catia", "nx", "solidworks", "ansys", "comsol", "matlab",
            "simulacion", "research scientist", "materials", "material"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Product Engineering/Design | C-Level",
        "keywords": [
            "diseno de producto", "product design", "industrial design",
            "disenador industrial", "ingenieria de producto", "product engineer",
            "cad", "cae", "autocad", "solidworks", "catia", "creo", "inventor",
            "fmea", "dfm", "dfa", "dfmea", "bom", "render", "maquetas",
            "prototipado", "ergonomia", "automotive", "construction"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Logistics (Transport/Freight)",
        "keywords": [
            "logistica", "transporte", "freight", "flete", "envios", "despacho",
            "expedicion", "distribucion", "3pl", "4pl", "courier", "logistico",
            "flota", "ruteo", "ultima milla", "last mile", "importaciones", "logisticas",
            "exportaciones", "aduanas", "customs", "forwarder", "log?stica", "logistic",
            "incoterms", "tms", "wms", "almacen", "bodega", "warehouse", "inventario",
            "cross docking", "construccion", "retail", "obra", "fabrica", "obras", "procesos",
            "logisitica", "logistics", "inventarios", "inventories", "transportation", "logi?stica"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Manufacturing",
        "keywords": [
            "manufactura", "fabricacion", "produccion", "co-manufacturing", "textil", "taller",
            "jefe de produccion", "supervisor de produccion", "ingeniero de procesos",
            "process engineer", "manufacturing", "six sigma", "oee", "kaizen",
            "gemba", "tpm", "5s", "apqp", "ppap", "spc", "quality", "calidad", "mes",
            "sap pp", "maquinaria", "operario", "shop floor", "empaque", "mining"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Customer Service",
        "keywords": [
            "cliente", "customer", "soporte", "postventa", "call center",
            "contact center", "mesa de ayuda", "helpdesk", "cx", "nps", "sac",
            "quejas", "reclamos", "retencion", "omnicanal", "omnichannel", "clientes",
            "client", "clients", "costumer"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Marketing",
        "keywords": [
            "marketing", "marca", "brand", "branding", "mercado",
            "market research", "insights", "posicionamiento", "atl", "btl",
            "comunicacion", "campanas", "patrocinios", "lanzamiento", "cmo", 
            "mkt", "mercadeo", "product", "producto", "marketin", "mercadotecnia",
            "publicidad", "creative"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Sales",
        "keywords": [
            "ventas", "comercial", "sales", "acount", "account", "accounts"
            "kae", "preventa", "presales", "posventa", "territory", "cuota", "quota", 
            "pipeline", "crm", "salesforce", "zoho", "hubspot", "negociacion", "hunter", 
            "farmer", "chief revenue officer", "cro", "chief sales officer", "cso", "commercial",
            "comercializacion", "demanda", "cuentas", "cuenta", "comercio"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Digital Transformation & Innovation",
        "keywords": [
            "digital", "innovacion", "automatizacion", "innovation",
            "internet de las cosas", "iot", "chief digital officer", "blockchain",
            "ai", "ia", "digitales", "transformation", "web", "network", "transformacion", "inteligencia artificial"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Audit",
        "keywords": [
            "auditor", "auditoria", "audit", "auditing", "auditora", "sox", 
            "control de calidad", "revisor fiscal", "revision fiscal", "interventoria",
            "compliance", "cumplimiento", "inspection", "inspeccion", "verificacion", "supervision",
            "monitoreo", "contralor", "contraloria", "contraloria", "control interno", "cro"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Compliance/Risk",
        "keywords": [
            "riesgos", "riesgo", "compliance", "cumplimiento", "regulatory", "Chief Compliance Officer", 
            "cco", "Chief Risk Officer", "cro", "risk", "etica", "regulacion", "regulaciones" "fraude",
            "cumplimient", "complaince", "compliancee", "cumpliminto"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Finance/Accounting",
        "keywords": [
            "finanzas", "contabilidad", "accounting", "tesoreria", "cobranza", "pagos", "gastos", "presupuestal",
            "cobrar", "facturacion", "impuestos", "tax", "cobros", "finance", "fianzas", "accounting",
            "financiero", "finance", "presupuesto", "budget", "financiera", "costos", "cost", "cfo",
            "presupuestos", "contable", "contador", "contadora", "tesorera", "liquidaciones", "fiscal", "valuaciones",
            "tributario", "credito", "creditos", "cobranzas", "tesoreria", "pagar", "fianza", "financial",
            "portfolio", "patrimonial", "patrimonio", "credit", "treasury", "revenue", "bank", "ingresos",
            "bursatil", "bustatil", "licitacion", "licitaciones", "cartera", "taxes", "financiamiento",
            "microfinanzas", "activo", "estadistica", "financieros", "tributaria", "tributario", "tributarias",
            "tributarios", "business unit", "tesorero", "investment", "accountant", "contabilida", "bancarias",
            "tesorer?a", "tesorer?a?", "inversiones", "inversion", "contaduria", "fondos", "expenses", "finanza", "portofolio"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Business Development",
        "keywords": [
            "desarrollo de negocios", "business", "bd", "bdm", "alianzas",
            "partnerships", "channel", "canales", "ecosistema", "corporate development",
            "expansion", "apertura de mercados", "go to market", "gtm", "licenciamiento",
            "estrategia comercial", "strategic partnerships", "hunter"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Media/Communications/PR",
        "keywords": [
            "comunicaciones", "comunicacion", "comunication", "comunications", "prensa", "press", "comunicadora",
            "comunicador", "public relations", "relacions publicas", "pr", "communications","communication", "journalist",
            "reporter", "periodista", "columnista", "news", "telecomunicaciones", "telecomunicacion"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Chief Executive Officer",
        "keywords": [
            "ceo", "general", "dueño", "propietario", "managing", "fundador", "executive", "area", "excecutive", "chairman",
            "presidente", "cofundador", "founder", "owner", "head", "ejecutiva", "sucursal", "branch", "distrito",
            "vp", "vicepresidente", "board", "country", "administrativo", "admnistrativo", "administrativa", "exec", "sede",
            "ejecutivo", "socio", "direccion general", "executive", "geral", "senior", "superintendente", "administraci?n",
            "founding", "fundadora", "regional", "administrative", "territorio", "team", "equipo", "region", "pais", "city", "zona", "vicepresident"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Shared Services",
        "keywords": [
            "servicios compartidos", "shared services", "ssc", "gbs", "global business services", 
            "centro de servicios", "captive", "bpo", "outsourcing", "torre de proceso", "ptp", "p2p", 
            "procure to pay", "otc", "order to cash", "rtr", "record to report", "sla", "kpi", "transicion", "migracion de procesos"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "SAP Competency Center",
        "keywords": [
            "sap", "centro de excelencia", "coe", "competency center", "abap", "s/4hana", "solman", "ariba", "ibp", "apo", "erp", "b1"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Asset Management (MRO)",
        "keywords": [
            "gestion de activos", "asset management", "mro", "mantenimiento", "repuestos", 
            "spares", "eam", "enterprise asset management", "apm", 
            "asset performance", "cmms", "maximo", "infor eam", "sap pm", "fiabilidad", "reliability", 
            "rcm", "inspeccion", "predictivo", "condition monitoring", "scada"
        ],
        "has_clevel_variant": False
    },
    {
        "base_name": "Legal",
        "keywords": [
            "legal", "abogado", "juridico", "corporate counsel", "general counsel", "lawyer",
            "gc", "contratos", "propiedad intelectual", "pi", "litigios", "regulatorio", "normatividad",
            "secretario corporativo", "notaria", "legales", "abogada", "normas", "legislacion", "normativa"
        ],
        "has_clevel_variant": True
    },
    {
        "base_name": "Sustainability",
        "keywords": [
            "sostenibilidad", "sustentabilidad", "esg", "medioambiente", "ambiental",
            "huella de carbono", "net zero", "carbono neutral", "rse", "csr", "ehs",
            "hse", "salud", "reportes esg", "gri", "sasb", "tcfd", "medioambiente",
            "economia circular", "energias renovables", "sustainability", "cso", "ambientes", "environmental", "ambiente"
        ],
        "has_clevel_variant": True
    }
]

# Generar variantes C-Level
persona_definitions = []
for base in persona_bases:
    persona_definitions.append({
        "name": base["base_name"],
        "keywords": base["keywords"],
        "requires_clevel": False
    })
    if base["has_clevel_variant"]:
        persona_definitions.append({
            "name": f"{base['base_name']} | C-Level",
            "keywords": base["keywords"],
            "requires_clevel": True
        })

# ============================================================
# 🔍 FUNCIONES DE CLASIFICACIÓN
# ============================================================

clevel_no = ['analista', 'asistente', 'ayudante', 'intern', "trainee", "pasante", "junior", "encargado", "consultor", "ingeniero", "supervisor",
             "reprsentante", "responsable", "analyst"]

clevel_si = ['manager', 'ceo', 'cfo', 'jefe', 'dueño', "gerente", "coo", "cpo", "cso", "chief", "owner", "presidente", "president", "founder", "vp",
             "vicepresident", "vice", "leader", "cmo", "chro", "cio", "cto", "cdo", "director", "vicepresidente", "socio", "cofounder", "co-founder",
             "coordinador", "lead", "cofundador", "fundador", "senior", "superintendente", "subgerente", "partner", "fundadora", "subdirector", "directora",
             "lider", "head", "chairman", "lider", "director", "jefa", "coordinadora", "coord", "gte", "LÃDER", "dir", "administrador", "administradora",
             "coordinator", "encargado", "encargada", "executive", "jefatura", "board", "direcciÃ³n", "subdirectora", "subdireccion", "hrbp", "subdir",
             "subgerencia", "executive"]


lob_fs = list({kw for b in persona_bases if b["base_name"] in ["Finance/Accounting", "Procurement/Purchasing"] for kw in b["keywords"]})

lob_hcm = list({kw for b in persona_bases if b["base_name"] in [
    "Human Resources | Compensation and Benefits", "Human Resources | Diversity and Inclusions", "Human Resources | Employee Experience", "Human Resources | Generalist",
    "Human Resources | HRIT", "Human Resources | Learning & Development", "Human Resources | Learning & Development","Human Resources | Payroll",
    "Human Resources | Talent Acquisition", "Human Resources | C-Level"] for kw in b["keywords"]})

lob_s4 = list({kw for b in persona_bases if b["base_name"] in [
    "Finance/Accounting", "Inform. Technology", "Data Protection, Privacy & Information Security", "Analytics and Data Warehousing",
    "Database, Data Management & Information","Digital Transformation & Innovation", "Operations | C-Level", "Chief Executive Officer"] for kw in b["keywords"]}| 
    {"ceo", "chief executive officer", "director general", "presidente", "senior manager", "superintendente"})

lob_btp = list({kw for b in persona_bases if b["base_name"] in [
    "Inform. Technology", "Data Protection, Privacy & Information Security", "Analytics and Data Warehousing",
    "Database, Data Management & Information","Digital Transformation & Innovation" ] for kw in b["keywords"]})

lob_cx = list({kw for b in persona_bases if b["base_name"] in [
    "Sales", "E-Commerce", "Marketing", "Digital Marketing", "Business Development", "Customer Service"] for kw in b["keywords"]})

lob_dsc = list({kw for b in persona_bases if b["base_name"] in [
    "Logistics (Transport/Freight)", "Product Engineering/Design | C-Level", "Supply Chain Planning", "Operations | C-Level",
    "Product Development / R&D / Engineering", "Plant Maintenance", "Sustaintability", "Manufacturing"] for kw in b["keywords"]})

def normalizar(texto):
    texto = str(texto).lower()
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def contains_word(text, words):
    if pd.isna(text):
        return False
    text = normalizar(text)
    for word in words:
        word_norm = normalizar(word)
        if re.search(rf'(?<!\w){word_norm}(?!\w)', text):
            return True
    return False

def get_clevel(title):
    if contains_word(title, clevel_si):
        return 'Si'
    elif contains_word(title, clevel_no):
        return 'No'
    else:
        return ''

def get_lob(title):
    if contains_word(title, lob_fs):
        return 'F&S'
    elif contains_word(title, lob_hcm):
        return 'HCM'
    elif contains_word(title, lob_btp):
        return 'BTP'
    elif contains_word(title, lob_dsc):
        return 'DSC'
    elif contains_word(title, lob_cx):
        return 'CX'
    elif contains_word(title, lob_s4):
        return 'S4'
    else:
        return ''

def get_s4(title):
    return 'Si' if contains_word(title, lob_s4) else 'No'

def get_persona_name(title, clevel):
    matches = []
    for persona in persona_definitions:
        if contains_word(title, persona["keywords"]):
            matches.append(persona)
    if not matches:
        return ''
    if clevel == 'Si':
        for persona in matches:
            if persona["requires_clevel"]:
                return persona["name"]
    for persona in matches:
        if not persona["requires_clevel"]:
            return persona["name"]
    return ''

# ============================================================
# 🧾 GENERACIÓN Y FORMATO FINAL
# ============================================================

df['Clevel'] = df['Titles'].apply(get_clevel)
df['LOB'] = df['Titles'].apply(get_lob)
df['Persona Name'] = df.apply(lambda row: get_persona_name(row['Titles'], row['Clevel']), axis=1)
df['S4?'] = df['Titles'].apply(get_s4)

df.loc[df['Persona Name'].isin(['Inform. Technology | C-Level', 'Inform. Technology', 'Chief Executive Officer']), 'S4?'] = 'Si'
df.loc[df['Persona Name'].str.contains('Inform. Technology', na=False), 'LOB'] = 'BTP'

df = df[['Key', 'Titles', 'Persona Name', 'Clevel', 'LOB', 'S4?']]
df.to_excel(output_file, index=False)

# ============================================================
# 🎨 FORMATO VISUAL DEL EXCEL
# ============================================================

wb = load_workbook(output_file)
ws = wb.active

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
center_align = Alignment(horizontal="center", vertical="center")
border_style = Border(
    left=Side(style='thin', color="CCCCCC"),
    right=Side(style='thin', color="CCCCCC"),
    top=Side(style='thin', color="CCCCCC"),
    bottom=Side(style='thin', color="CCCCCC")
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border_style

col_widths = {"A": 15, "B": 45, "C": 40, "D": 10, "E": 10}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = center_align
        cell.border = border_style

wb.save(output_file)
print(f"✅ Archivo generado correctamente con formato: {output_file}")
